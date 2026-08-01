import re
import unicodedata
from collections import defaultdict
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

source_path = Path('81Horizon-Atlas-complete-1000-destinations-master-workbook.xlsx')
out_path = Path('Duplicate-Review-Recommendations.xlsx')

wb_source = load_workbook(source_path, data_only=True, read_only=True)

destination_sheet = None
for ws in wb_source.worksheets:
    rows = list(ws.iter_rows(min_row=1, max_row=2, values_only=True))
    if not rows:
        continue
    headers = [re.sub(r'[^a-z0-9]+', '_', unicodedata.normalize('NFKC', str(cell)).strip().lower()).strip('_') for cell in rows[0]]
    if {'slug', 'city', 'country'} <= set(headers):
        destination_sheet = ws
        break

if destination_sheet is None:
    raise RuntimeError('Could not find destination sheet in workbook')

rows = list(destination_sheet.iter_rows(values_only=True))
headers = [re.sub(r'[^a-z0-9]+', '_', unicodedata.normalize('NFKC', str(cell)).strip().lower()).strip('_') for cell in rows[0]]

all_rows = []
for row_idx, row in enumerate(rows[1:], start=2):
    if not any(unicodedata.normalize('NFKC', str(cell)).strip() for cell in row):
        continue
    record = {}
    for idx, header in enumerate(headers):
        if idx < len(row):
            record[header] = unicodedata.normalize('NFKC', str(row[idx]))
    slug = (record.get('slug') or '').strip()
    if slug:
        all_rows.append((row_idx, slug, record))

by_slug = defaultdict(list)
for row_idx, slug, record in all_rows:
    by_slug[slug].append((row_idx, record))

out_wb = Workbook()
summary_ws = out_wb.active
summary_ws.title = 'Summary'
summary_headers = [
    'Slug', 'Destination Name', 'Country', 'Row A', 'Row B',
    'Duplicate Type', 'Recommended Master Row', 'Recommended Master Name',
    'Recommended Master Country', 'Fields to Merge', 'Reason', 'Recommended Corrected Slug'
]
summary_ws.append(summary_headers)

header_fill = PatternFill('solid', fgColor='D9EAF7')
header_font = Font(bold=True)
for cell in summary_ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(wrap_text=True)


def normalize_text(value):
    if value is None:
        return ''
    return unicodedata.normalize('NFKC', str(value)).strip()


def row_score(record):
    score = 0
    for field, value in record.items():
        text = normalize_text(value)
        if not text:
            continue
        weight = 2 if field in {'description', 'overview', 'climate', 'lifestyle', 'transportation', 'notes', 'research_notes', 'relocation_profile', 'cost_estimate', 'costs', 'safety', 'healthcare', 'airport'} else 1
        score += weight
    return score

sheet_count = 0
for index, (slug, entries) in enumerate(sorted(by_slug.items()), start=1):
    if len(entries) < 2:
        continue

    row_a_idx, row_a = entries[0]
    row_b_idx, row_b = entries[1]

    score_a = row_score(row_a)
    score_b = row_score(row_b)
    if score_b > score_a:
        master_row_idx, master_row, other_row_idx, other_row = row_b_idx, row_b, row_a_idx, row_a
    else:
        master_row_idx, master_row, other_row_idx, other_row = row_a_idx, row_a, row_b_idx, row_b

    if score_a == score_b and row_a_idx < row_b_idx:
        master_row_idx, master_row, other_row_idx, other_row = row_a_idx, row_a, row_b_idx, row_b

    city_a = normalize_text(row_a.get('city'))
    city_b = normalize_text(row_b.get('city'))
    country_a = normalize_text(row_a.get('country'))
    country_b = normalize_text(row_b.get('country'))
    same_destination = (city_a.lower() == city_b.lower() and country_a.lower() == country_b.lower())
    duplicate_type = 'same destination repeated' if same_destination else 'different destinations share the same slug'

    corrected_slug = ''
    if duplicate_type != 'same destination repeated':
        corrected_slug = f"{re.sub(r'[^a-z0-9]+', '-', (city_a or city_b).lower()).strip('-')}-{re.sub(r'[^a-z0-9]+', '-', (country_a or country_b).lower()).strip('-')}"

    fields_to_merge = []
    for header in headers:
        a_val = normalize_text(row_a.get(header))
        b_val = normalize_text(row_b.get(header))
        if not a_val or not b_val or a_val == b_val:
            continue
        fields_to_merge.append(header)

    reason = (
        f"Row {master_row_idx} was selected as the master because it contains the richest populated fields "
        f"for this destination ({score_a if master_row_idx == row_a_idx else score_b} vs {score_b if master_row_idx == row_a_idx else score_a}). "
        f"Row {other_row_idx} contributes the unique values that should be reviewed and merged into the master row."
    )

    sheet_name = f"dup_{index:02d}_{re.sub(r'[^a-z0-9]+', '_', slug.lower()).strip('_')[:24]}"
    sheet_name = sheet_name[:31]
    detail_ws = out_wb.create_sheet(title=sheet_name)
    sheet_count += 1
    detail_ws.append(['Field', 'Original Row A', 'Original Row B', 'Recommended Master Row', 'Merge Action'])
    for cell in detail_ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    detail_ws.append(['Slug', slug, slug, slug, 'keep'])
    detail_ws.append(['Duplicate Type', duplicate_type, duplicate_type, duplicate_type, 'keep'])
    detail_ws.append(['Destination Name', city_a, city_b, normalize_text(master_row.get('city')), 'keep'])
    detail_ws.append(['Country', country_a, country_b, normalize_text(master_row.get('country')), 'keep'])
    detail_ws.append(['Workbook Row A', row_a_idx, '', '', ''])
    detail_ws.append(['Workbook Row B', '', row_b_idx, '', ''])
    detail_ws.append(['Recommended Master Row', '', '', master_row_idx, ''])
    detail_ws.append(['Reason for Recommendation', '', '', reason, ''])
    detail_ws.append(['Fields to Merge', '', '', ', '.join(fields_to_merge), ''])
    detail_ws.append([])
    detail_ws.append(['Field', 'Original Row A Value', 'Original Row B Value', 'Recommended Master Value', 'Action'])
    for header in headers:
        a_val = normalize_text(row_a.get(header))
        b_val = normalize_text(row_b.get(header))
        if not a_val and not b_val:
            action = 'none'
            merged_val = ''
        elif not a_val:
            action = 'take row B value'
            merged_val = b_val
        elif not b_val:
            action = 'take row A value'
            merged_val = a_val
        elif a_val == b_val:
            action = 'same value'
            merged_val = a_val
        else:
            action = 'review and merge'
            merged_val = normalize_text(master_row.get(header))
        detail_ws.append([header, a_val, b_val, merged_val, action])
    for col_idx in range(1, detail_ws.max_column + 1):
        detail_ws.column_dimensions[get_column_letter(col_idx)].width = 22
    detail_ws.freeze_panes = 'A2'

    summary_ws.append([
        slug,
        normalize_text(master_row.get('city')),
        normalize_text(master_row.get('country')),
        row_a_idx,
        row_b_idx,
        duplicate_type,
        master_row_idx,
        normalize_text(master_row.get('city')),
        normalize_text(master_row.get('country')),
        ', '.join(fields_to_merge),
        reason,
        corrected_slug,
    ])

for row in summary_ws.iter_rows(min_row=2, max_col=summary_ws.max_column):
    for cell in row:
        cell.alignment = Alignment(wrap_text=True, vertical='top')

out_wb.save(out_path)
print(f'Created {out_path}')
print(f'Processed {len(by_slug)} unique slugs and wrote {sheet_count} duplicate review sheets')
