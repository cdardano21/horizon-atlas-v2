from __future__ import annotations

import argparse
import json
import re
import unicodedata
from pathlib import Path
from typing import Any

import openpyxl

REPO_ROOT = Path(__file__).resolve().parent.parent
DESTINATIONS_PATH = REPO_ROOT / "app/lib/destinations.ts"
DEFAULT_WORKBOOK = "81Horizon-Atlas-CLEAN.xlsx"
REQUIRED_FIELDS = ("slug", "city", "country", "description", "overview", "climate", "lifestyle", "transportation")


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return unicodedata.normalize("NFKC", value)
    return str(value)


def normalize_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "_", normalize_text(value).strip().lower()).strip("_")


def normalize_emoji(value: Any) -> str:
    text = normalize_text(value).strip()
    if not text:
        return "🌍"
    try:
        return text.encode("latin1", "ignore").decode("utf-8", "ignore")
    except Exception:
        return text


def parse_tags(value: Any) -> list[str]:
    text = normalize_text(value).strip()
    if not text:
        return []
    parts = [part.strip() for part in re.split(r"[;,]", text) if part.strip()]
    return [part for part in parts if part][:8]


def get_field(record: dict[str, str], *names: str) -> str:
    for name in names:
        value = record.get(name, "")
        if value:
            return normalize_text(value)
    return ""


def build_entry(record: dict[str, str]) -> tuple[dict[str, Any] | None, list[str]]:
    slug = get_field(record, "slug").strip()
    city = get_field(record, "city").strip()
    country = get_field(record, "country").strip()
    description = get_field(record, "description").strip()
    overview = get_field(record, "overview").strip()
    climate = get_field(record, "climate").strip()
    lifestyle = get_field(record, "lifestyle").strip()
    transportation = get_field(record, "transportation").strip()

    missing_fields = [field for field, value in {
        "slug": slug,
        "city": city,
        "country": country,
        "description": description,
        "overview": overview,
        "climate": climate,
        "lifestyle": lifestyle,
        "transportation": transportation,
    }.items() if not value]
    if missing_fields:
        return None, missing_fields

    match_text = get_field(record, "match").strip()
    try:
        match_value = float(match_text.replace(",", "")) if match_text else 0.0
    except ValueError:
        match_value = 0.0

    return {
        "slug": slug,
        "city": city,
        "country": country,
        "emoji": normalize_emoji(get_field(record, "emoji")),
        "match": match_value,
        "description": description,
        "overview": overview,
        "climate": climate,
        "lifestyle": lifestyle,
        "transportation": transportation,
        "caption": f"{city}, {country}",
        "tags": parse_tags(get_field(record, "tags")),
    }, []


def find_destination_sheet(workbook: openpyxl.Workbook) -> openpyxl.worksheet.worksheet.Worksheet:
    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(min_row=1, max_row=2, values_only=True))
        if not rows:
            continue
        headers = [normalize_header(cell) for cell in rows[0]]
        if {"slug", "city", "country"}.issubset(set(headers)):
            return sheet
    raise ValueError("Could not locate a destination worksheet with slug/city/country headers.")


def read_workbook_entries(workbook_path: Path) -> tuple[list[dict[str, Any]], list[str], dict[str, list[str]]]:
    workbook = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)
    sheet = find_destination_sheet(workbook)
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return [], [], {}

    raw_headers = [normalize_text(cell) for cell in rows[0]]
    header_names = [normalize_header(cell) for cell in raw_headers]
    entries: list[dict[str, Any]] = []
    duplicates: list[str] = []
    missing_by_slug: dict[str, list[str]] = {}
    seen_slugs: set[str] = set()

    for row in rows[1:]:
        if not any(normalize_text(cell) for cell in row):
            continue

        record = {
            header_names[index]: normalize_text(value)
            for index, value in enumerate(row)
            if index < len(header_names)
        }
        entry, missing_fields = build_entry(record)
        if entry is None:
            slug = get_field(record, "slug")
            missing_by_slug[slug or "<missing-slug>"] = missing_fields
            continue

        if entry["slug"] in seen_slugs:
            duplicates.append(entry["slug"])
            continue

        seen_slugs.add(entry["slug"])
        entries.append(entry)

    return entries, duplicates, missing_by_slug


def write_destinations_catalog(entries: list[dict[str, Any]]) -> None:
    lines: list[str] = ["export const destinations: Destination[] = ["]
    for entry in entries:
        lines.append("  {")
        lines.append(f'    slug: {json.dumps(entry["slug"], ensure_ascii=False)},')
        lines.append(f'    city: {json.dumps(entry["city"], ensure_ascii=False)},')
        lines.append(f'    country: {json.dumps(entry["country"], ensure_ascii=False)},')
        lines.append(f'    emoji: {json.dumps(entry["emoji"], ensure_ascii=False)},')
        lines.append(f'    match: {entry["match"]},')
        lines.append(f'    description: {json.dumps(entry["description"], ensure_ascii=False)},')
        lines.append(f'    overview: {json.dumps(entry["overview"], ensure_ascii=False)},')
        lines.append(f'    climate: {json.dumps(entry["climate"], ensure_ascii=False)},')
        lines.append(f'    lifestyle: {json.dumps(entry["lifestyle"], ensure_ascii=False)},')
        lines.append(f'    transportation: {json.dumps(entry["transportation"], ensure_ascii=False)},')
        lines.append("    images: [")
        lines.append("      {")
        lines.append('        src: "",')
        lines.append(f'        alt: {json.dumps(entry["city"] + " city view", ensure_ascii=False)},')
        lines.append(f'        caption: {json.dumps(entry["caption"], ensure_ascii=False)},')
        lines.append("      },")
        lines.append("    ],")
        lines.append(f'    tags: [{", ".join(json.dumps(tag, ensure_ascii=False) for tag in entry["tags"])}],')
        lines.append("  },")
    lines.append("];")
    lines.append("")
    lines.append("export const LAUNCH_CATALOG_SIZE = destinations.length;")

    destinations_text = DESTINATIONS_PATH.read_text(encoding="utf-8")
    pattern = re.compile(r"export const destinations: Destination\[\] = \[(.*?)\n\];\n\nexport const LAUNCH_CATALOG_SIZE = destinations.length;", re.S)
    updated = pattern.sub("\n".join(lines), destinations_text, count=1)
    DESTINATIONS_PATH.write_text(updated, encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser(description="Import destination rows from the master workbook into the app catalog.")
    parser.add_argument("workbook", nargs="?", default=DEFAULT_WORKBOOK)
    args = parser.parse_args()

    workbook_path = REPO_ROOT / args.workbook if not Path(args.workbook).is_absolute() else Path(args.workbook)
    entries, duplicates, missing_by_slug = read_workbook_entries(workbook_path)
    write_destinations_catalog(entries)

    print(f"Imported {len(entries)} destinations from {workbook_path.name}.")
    print(f"Duplicate slugs: {len(duplicates)}")
    print(f"Missing required fields: {len(missing_by_slug)}")
    if duplicates:
        print("Duplicate slug examples:", ", ".join(duplicates[:10]))
    if missing_by_slug:
        print("Missing field examples:")
        for slug, fields in list(missing_by_slug.items())[:10]:
            print(f"  - {slug}: {', '.join(fields)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
