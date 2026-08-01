from __future__ import annotations

import re
from pathlib import Path
from typing import List, Dict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


REPO_ROOT = Path(__file__).resolve().parent.parent
SOURCES_PATH = REPO_ROOT / "app/lib/destinations.ts"
OUTPUT_PATH = REPO_ROOT / "destinations-export.xlsx"


def extract_array_body(source: str, array_name: str) -> str:
    pattern = re.compile(rf"export const {re.escape(array_name)}: Destination\[\] = \[(.*?)\n\];", re.S)
    match = pattern.search(source)
    if not match:
        raise ValueError(f"Could not find destinations array in {SOURCES_PATH}")
    return match.group(1)


def parse_top_level_objects(body: str) -> List[str]:
    objects: List[str] = []
    depth = 0
    in_string = False
    escape = False
    in_line_comment = False
    in_block_comment = False
    started = False
    current: List[str] = []

    i = 0
    while i < len(body):
        ch = body[i]
        nxt = body[i + 1] if i + 1 < len(body) else ""

        if in_line_comment:
            if ch == "\n":
                in_line_comment = False
            current.append(ch)
            i += 1
            continue

        if in_block_comment:
            if ch == "*" and nxt == "/":
                in_block_comment = False
                current.append("/")
                i += 2
                continue
            current.append(ch)
            i += 1
            continue

        if in_string:
            current.append(ch)
            if escape:
                escape = False
            elif ch == "\\":
                escape = True
            elif ch == '"':
                in_string = False
            i += 1
            continue

        if ch == "/" and nxt == "/":
            in_line_comment = True
            current.append(ch)
            current.append(nxt)
            i += 2
            continue

        if ch == "/" and nxt == "*":
            in_block_comment = True
            current.append(ch)
            current.append(nxt)
            i += 2
            continue

        if ch == '"':
            in_string = True
            current.append(ch)
            i += 1
            continue

        if ch == "{":
            if not started:
                started = True
                depth = 1
                current.append(ch)
                i += 1
                continue
            depth += 1
            current.append(ch)
            i += 1
            continue

        if ch == "}":
            if started:
                depth -= 1
                current.append(ch)
                if depth == 0:
                    objects.append("".join(current).strip())
                    current = []
                    started = False
                i += 1
                continue

        if started:
            current.append(ch)
        i += 1

    return objects


def extract_string_property(text: str, property_name: str) -> str:
    pattern = re.compile(rf'\b{re.escape(property_name)}\s*:\s*"((?:\\.|[^"\\])*)"', re.S)
    match = pattern.search(text)
    if not match:
        return ""
    return match.group(1).encode("utf-8").decode("unicode_escape")


def extract_int_property(text: str, property_name: str) -> str:
    pattern = re.compile(rf'\b{re.escape(property_name)}\s*:\s*(\d+)', re.S)
    match = pattern.search(text)
    return match.group(1) if match else ""


def extract_tags(text: str) -> str:
    match = re.search(r'\btags\s*:\s*\[(.*?)\]', text, re.S)
    if not match:
        return ""
    values = re.findall(r'"([^\"]+)"', match.group(1))
    return "; ".join(values)


def build_records(objects: List[str]) -> List[Dict[str, object]]:
    records: List[Dict[str, object]] = []
    for obj in objects:
        slug = extract_string_property(obj, "slug")
        if not slug:
            continue
        records.append(
            {
                "slug": slug,
                "city": extract_string_property(obj, "city"),
                "country": extract_string_property(obj, "country"),
                "emoji": extract_string_property(obj, "emoji"),
                "match": extract_int_property(obj, "match"),
                "description": extract_string_property(obj, "description"),
                "overview": extract_string_property(obj, "overview"),
                "climate": extract_string_property(obj, "climate"),
                "lifestyle": extract_string_property(obj, "lifestyle"),
                "transportation": extract_string_property(obj, "transportation"),
                "tags": extract_tags(obj),
            }
        )
    return records


def write_excel(records: List[Dict[str, object]], output_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "destinations"

    headers = [
        "slug",
        "city",
        "country",
        "emoji",
        "match",
        "description",
        "overview",
        "climate",
        "lifestyle",
        "transportation",
        "tags",
    ]
    sheet.append(headers)

    for record in records:
        sheet.append(
            [
                record.get("slug", ""),
                record.get("city", ""),
                record.get("country", ""),
                record.get("emoji", ""),
                record.get("match", ""),
                record.get("description", ""),
                record.get("overview", ""),
                record.get("climate", ""),
                record.get("lifestyle", ""),
                record.get("transportation", ""),
                record.get("tags", ""),
            ]
        )

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max_length + 2, 80)

    workbook.save(output_path)


def main() -> None:
    source = SOURCES_PATH.read_text(encoding="utf-8")
    body = extract_array_body(source, "destinations")
    objects = parse_top_level_objects(body)
    records = build_records(objects)
    write_excel(records, OUTPUT_PATH)
    print(f"Wrote {len(records)} destinations to {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
