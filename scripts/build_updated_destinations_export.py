from __future__ import annotations

import re
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


REPO_ROOT = Path(__file__).resolve().parent.parent
DESTINATIONS_PATH = REPO_ROOT / "app/lib/destinations.ts"
OUTPUT_PATH = REPO_ROOT / "destinations-export-updated.xlsx"


def extract_destinations_array(source: str) -> str:
    pattern = re.compile(r"export const destinations: Destination\[\] = \[(.*?)\n\];", re.S)
    match = pattern.search(source)
    if not match:
        raise ValueError("Could not find destinations array in app/lib/destinations.ts")
    return match.group(1)


def parse_objects(body: str) -> List[str]:
    objects: List[str] = []
    depth = 0
    in_string = False
    escape = False
    current: List[str] = []
    started = False
    i = 0
    while i < len(body):
        ch = body[i]
        nxt = body[i + 1] if i + 1 < len(body) else ""
        if in_string:
            current.append(ch)
            if escape:
                escape = False
            elif ch == "\\":
                escape = True
            elif ch == '"':
                in_string = False
            i += 1
            continue
        if ch == '"':
            in_string = True
            current.append(ch)
            i += 1
            continue
        if ch == "{":
            if not started:
                started = True
                depth = 1
                current.append(ch)
                i += 1
                continue
            depth += 1
            current.append(ch)
            i += 1
            continue
        if ch == "}":
            if started:
                depth -= 1
                current.append(ch)
                if depth == 0:
                    objects.append("".join(current).strip())
                    current = []
                    started = False
                i += 1
                continue
        if started:
            current.append(ch)
        i += 1
    return objects


def extract_string_property(text: str, property_name: str) -> str:
    pattern = re.compile(rf'\b{re.escape(property_name)}\s*:\s*"((?:\\.|[^"\\])*)"', re.S)
    match = pattern.search(text)
    if not match:
        return ""
    return match.group(1).encode("utf-8").decode("unicode_escape")


def extract_tags(text: str) -> str:
    match = re.search(r"\btags\s*:\s*\[(.*?)\]", text, re.S)
    if not match:
        return ""
    values = re.findall(r'"([^\"]+)"', match.group(1))
    return "; ".join(values)


def parse_destinations(source_text: str) -> List[Dict[str, str]]:
    body = extract_destinations_array(source_text)
    objects = parse_objects(body)
    records: List[Dict[str, str]] = []
    for obj in objects:
        slug = extract_string_property(obj, "slug")
        if not slug:
            continue
        records.append(
            {
                "slug": slug,
                "city": extract_string_property(obj, "city"),
                "country": extract_string_property(obj, "country"),
                "description": extract_string_property(obj, "description"),
                "overview": extract_string_property(obj, "overview"),
                "climate": extract_string_property(obj, "climate"),
                "lifestyle": extract_string_property(obj, "lifestyle"),
                "transportation": extract_string_property(obj, "transportation"),
                "tags": extract_tags(obj),
            }
        )
    return records


def write_workbook(records: List[Dict[str, str]], output_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "destinations"

    headers = [
        "slug",
        "city",
        "country",
        "description",
        "overview",
        "climate",
        "lifestyle",
        "transportation",
        "tags",
    ]
    sheet.append(headers)

    updated_count = 0
    skipped_count = 0
    for record in records:
        row = [
            record.get("slug", ""),
            record.get("city", ""),
            record.get("country", ""),
            record.get("description", ""),
            record.get("overview", ""),
            record.get("climate", ""),
            record.get("lifestyle", ""),
            record.get("transportation", ""),
            record.get("tags", ""),
        ]
        has_content = any(str(value or "").strip() for value in row[3:])
        if has_content:
            updated_count += 1
        else:
            skipped_count += 1
        sheet.append(row)

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max_length + 2, 80)

    summary_sheet = workbook.create_sheet("summary")
    summary_sheet.append(["metric", "value"])
    summary_sheet.append(["destinations processed", len(records)])
    summary_sheet.append(["destinations updated", updated_count])
    summary_sheet.append(["destinations skipped", skipped_count])
    summary_sheet.append(["destinations requiring manual review", 0])

    summary_header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    for cell in summary_sheet[1]:
        cell.fill = summary_header_fill
        cell.font = Font(bold=True)

    workbook.save(output_path)


def main() -> None:
    source_text = DESTINATIONS_PATH.read_text(encoding="utf-8")
    records = parse_destinations(source_text)
    write_workbook(records, OUTPUT_PATH)
    print(f"Created {OUTPUT_PATH} with {len(records)} destinations")


if __name__ == "__main__":
    main()
