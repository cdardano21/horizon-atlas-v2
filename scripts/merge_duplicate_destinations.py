from __future__ import annotations

import re
import unicodedata
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

ROOT = Path(__file__).resolve().parent.parent
SOURCE_WORKBOOK = ROOT / "81Horizon-Atlas-complete-1000-destinations-master-workbook.xlsx"
REVIEW_WORKBOOK = ROOT / "Duplicate-Review-Recommendations.xlsx"
OUTPUT_WORKBOOK = ROOT / "81Horizon-Atlas-CLEAN.xlsx"


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return unicodedata.normalize("NFKC", str(value)).strip()


def normalize_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "_", normalize_text(value).lower()).strip("_")


def is_list_like(field: str) -> bool:
    return field in {"tags", "tag", "keywords", "themes", "pros", "cons", "benefits"}


def merge_field(master_value: str, duplicate_value: str, field: str) -> str:
    master = normalize_text(master_value)
    duplicate = normalize_text(duplicate_value)

    if not master and not duplicate:
        return ""
    if not master:
        return duplicate
    if not duplicate:
        return master
    if master == duplicate:
        return master

    if is_list_like(field):
        items = []
        for raw in [master, duplicate]:
            for part in re.split(r"[;|,]", raw):
                item = part.strip()
                if item:
                    items.append(item)
        seen: set[str] = set()
        merged_parts: list[str] = []
        for item in items:
            if item.lower() not in seen:
                seen.add(item.lower())
                merged_parts.append(item)
        return "; ".join(merged_parts)

    if field in {"description", "overview", "climate", "lifestyle", "transportation", "notes", "research_notes", "relocation_profile", "summary", "story", "tagline"}:
        # Prefer the longer, richer version but preserve the other info if it is meaningfully different.
        if len(master) >= len(duplicate):
            combined = master
            if duplicate not in master:
                combined = f"{master} | {duplicate}"
            return combined
        combined = duplicate
        if master not in duplicate:
            combined = f"{duplicate} | {master}"
        return combined

    if len(master) >= len(duplicate):
        return master
    return duplicate


def read_review_recommendations(path: Path) -> list[dict[str, Any]]:
    review_wb = load_workbook(path, data_only=True, read_only=True)
    summary_ws = review_wb["Summary"]
    rows = list(summary_ws.iter_rows(values_only=True))
    if not rows:
        return []
    recommendations: list[dict[str, Any]] = []
    headers = [normalize_header(cell) for cell in rows[0]]
    for row in rows[1:]:
        if not any(normalize_text(cell) for cell in row):
            continue
        record = {header: normalize_text(value) for header, value in zip(headers, row)}
        if not record.get("slug"):
            continue
        recommendations.append(record)
    return recommendations


def find_destination_sheet(workbook: Workbook) -> Worksheet:
    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(min_row=1, max_row=2, values_only=True))
        if not rows:
            continue
        headers = [normalize_header(cell) for cell in rows[0]]
        if {"slug", "city", "country"}.issubset(set(headers)):
            return sheet
    raise RuntimeError("Could not locate destination sheet")


def build_clean_workbook() -> None:
    source_wb = load_workbook(SOURCE_WORKBOOK, data_only=False)
    target_wb = Workbook()

    # Copy sheets from source workbook to target workbook, preserving names and content.
    for sheet in source_wb.worksheets:
        new_sheet = target_wb.create_sheet(title=sheet.title)
        for row in sheet.iter_rows(values_only=False):
            new_sheet.append([cell.value for cell in row])

    # Remove the default empty sheet created by Workbook().
    if "Sheet" in target_wb.sheetnames:
        del target_wb["Sheet"]

    target_sheet = find_destination_sheet(target_wb)
    rows = list(target_sheet.iter_rows(values_only=True))
    headers = [normalize_header(cell) for cell in rows[0]]
    header_index = {header: idx for idx, header in enumerate(headers)}

    recommendations = read_review_recommendations(REVIEW_WORKBOOK)
    duplicate_groups: list[dict[str, Any]] = []
    for recommendation in recommendations:
        if not recommendation.get("slug"):
            continue
        row_a = recommendation.get("row_a", "")
        row_b = recommendation.get("row_b", "")
        master_row = recommendation.get("recommended_master_row", "")
        if row_a and row_b and master_row:
            duplicate_groups.append(
                {
                    "slug": recommendation["slug"],
                    "row_a": int(row_a),
                    "row_b": int(row_b),
                    "master_row": int(master_row),
                }
            )

    # Merge values into the master row, then remove duplicate rows in reverse order.
    rows_to_delete: list[int] = []
    for group in duplicate_groups:
        master_row = group["master_row"]
        duplicate_row = group["row_b"] if group["row_b"] != master_row else group["row_a"]
        rows_to_delete.append(duplicate_row)

        for col_idx, header in enumerate(headers, start=1):
            master_value = target_sheet.cell(row=master_row, column=col_idx).value
            duplicate_value = target_sheet.cell(row=duplicate_row, column=col_idx).value
            merged_value = merge_field(master_value if master_value is not None else "", duplicate_value if duplicate_value is not None else "", header)
            target_sheet.cell(row=master_row, column=col_idx, value=merged_value)

    # Preserve the recommended master row values and remove only duplicate rows.
    for row_idx in sorted(set(rows_to_delete), reverse=True):
        target_sheet.delete_rows(row_idx)

    # Ensure the workbook is saved as a clean copy and not overwrite the source.
    target_wb.save(OUTPUT_WORKBOOK)
    print(f"Created {OUTPUT_WORKBOOK}")
    print(f"Merged {len(duplicate_groups)} duplicate groups")


if __name__ == "__main__":
    build_clean_workbook()
